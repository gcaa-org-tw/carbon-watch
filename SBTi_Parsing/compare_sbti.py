"""Compare greentrade (today) vs upstream scrape (Nov 2025) vs carbon-watch site values."""

import csv
import json

SP = "/private/tmp/claude-501/-Users-jinsoon/60dece08-921d-4986-9594-c76b553e113e/scratchpad"
CW = "/Users/jinsoon/Work/Projects/thaubing/carbon-watch"

EN2ZH = {
    "Aligned to 1.5°C": "符合1.5°C目標",
    "Well-below 2°C": "遠低於2°C",
    "Committed to set SBT target": "承諾設定科學基礎目標",
    "X": "X",
}


def norm(s):
    return (s or "").replace("臺", "台").strip()


# --- inputs ---
gt = json.load(open(f"{SP}/greentrade_companies.json", encoding="utf-8"))[
    "records"
]  # 211 today
up = json.load(open(f"{SP}/upstream_companies.json", encoding="utf-8"))  # 191 rows A:E
raw = open(f"{SP}/listed_raw.txt", encoding="utf-8").read()
listed = json.loads(raw[raw.find("{") :])["values"][
    1:
]  # A統編 B市場 C日期 D代號 E全名 F簡稱
raw = open(f"{SP}/unlisted_raw.txt", encoding="utf-8").read()
unlisted = json.loads(raw[raw.find("{") :])["values"][1:]  # A全名 B簡稱 C統編
dahu = list(
    csv.DictReader(open(f"{CW}/raw-data/排碳大戶表_Data.csv", encoding="utf-8-sig"))
)  # 287

# lookup maps
listed_exact = {
    r[4]: r[0] for r in listed if len(r) > 4
}  # 全名 -> 統編 (exact, formula behavior)
listed_norm = {norm(r[4]): r[0] for r in listed if len(r) > 4}  # normalized
unlisted_norm = {norm(r[0]): r[2] for r in unlisted if len(r) > 2}
dahu_by_ubn = {r["統一編號"].strip(): r for r in dahu}
up_by_name = {norm(r[0]): r for r in up}
gt_by_name = {norm(r["公司名稱"]): r for r in gt}

print("=" * 60)
print("A. greentrade 今日 (211) vs 上游爬蟲表 2025-11 (191)")
added = [n for n in gt_by_name if n not in up_by_name]
removed = [n for n in up_by_name if n not in gt_by_name]
changed = []
for n, g in gt_by_name.items():
    u = up_by_name.get(n)
    if u is None:
        continue
    new_zh = EN2ZH[g["SBTi"]]
    old_zh = (u[1] if len(u) > 1 else "").strip()
    if new_zh != old_zh:
        changed.append((n, old_zh, new_zh))
print(f"新增 {len(added)}: {added}")
print(f"消失 {len(removed)}: {removed}")
print(f"SBTi 狀態變更 {len(changed)}:")
for n, o, nw in changed:
    print(f"  {n}: {o} -> {nw}")

print("=" * 60)
print("B. 模擬統編比對（照現行公式邏輯：上市櫃全名 exact match）")
sim = []
for g in gt:
    name = g["公司名稱"].strip()
    ubn = listed_exact.get(name)
    via = "上市櫃exact"
    if ubn is None:
        ubn = listed_norm.get(norm(name))
        via = "上市櫃臺台正規化後" if ubn else None
    if ubn is None:
        u2 = unlisted_norm.get(norm(name))
        if u2:
            ubn, via = u2, "非上市櫃（現行公式比不到，修好才會中）"
    sim.append({**g, "統編": ubn, "via": via})
no_ubn = [s["公司名稱"] for s in sim if s["統編"] is None]
norm_only = [s["公司名稱"] for s in sim if s["via"] and "正規化" in s["via"]]
unlisted_hits = [s["公司名稱"] for s in sim if s["via"] and "非上市櫃" in s["via"]]
print(
    f"查無統編（不影響排碳大戶欄，僅代表非上市櫃又不在名單）{len(no_ubn)}: {no_ubn[:15]}{'...' if len(no_ubn) > 15 else ''}"
)
print(f"僅靠臺→台正規化才中（現行 exact 公式會漏）{len(norm_only)}: {norm_only}")
print(
    f"落在非上市櫃名單（受 H 欄 fallback 公式 bug 影響）{len(unlisted_hits)}: {unlisted_hits}"
)

print("=" * 60)
print("C. 對排碳大戶 287 家的站上影響（greentrade 新值 vs 網站現值）")
diff = []
for s in sim:
    if s["統編"] and s["統編"] in dahu_by_ubn:
        d = dahu_by_ubn[s["統編"]]
        cur = d["中期目標是否取得SBT認證"].strip()
        new = EN2ZH[s["SBTi"]]
        mark = "SAME" if cur == new else "DIFF"
        diff.append((mark, d["公司簡稱"], s["公司名稱"], cur, new, s["via"]))
matched = len(diff)
diffs = [x for x in diff if x[0] == "DIFF"]
print(f"greentrade∩排碳大戶: {matched} 家；其中站上值會變 {len(diffs)} 家:")
for _, abbr, full, cur, new, via in diffs:
    print(f"  {abbr}（{full}）: 「{cur}」->「{new}」  [{via}]")
# 反向：站上有 SBTi 值但今日 greentrade 沒有的公司
gt_ubns = {s["統編"] for s in sim if s["統編"]}
for d in dahu:
    cur = d["中期目標是否取得SBT認證"].strip()
    if cur not in ("無", "", "X") and d["統一編號"].strip() not in gt_ubns:
        print(f"  反向缺席: {d['公司簡稱']} 站上「{cur}」但今日 greentrade 無此公司")

print("=" * 60)
print("D. qyke xlsx（6/3 快照, 199 家）vs 今日 greentrade (211)")
from openpyxl import load_workbook

wb = load_workbook(f"{SP}/greentrade_snapshot.xlsx", data_only=True)
ws = wb["企業名單"]
xr = [r for r in ws.iter_rows(values_only=True) if any(v not in (None, "") for v in r)]
hdr = [str(h) for h in xr[0]]
i_name, i_dahu = hdr.index("公司名稱"), hdr.index("排碳大戶")
xlsx_names = {norm(str(r[i_name])) for r in xr[1:]}
self_eq = sum(1 for r in xr[1:] if str(r[i_dahu]).strip() == str(r[i_name]).strip())
print(
    f"xlsx 排碳大戶欄 = 自身公司名: {self_eq}/{len(xr) - 1}（其餘值樣本: "
    f"{[str(r[i_dahu])[:20] for r in xr[1:] if str(r[i_dahu]).strip() != str(r[i_name]).strip()][:5]}）"
)
ws300 = wb["所有排碳大戶_2024_300"]
lst300 = {
    norm(str(r[0]))
    for r in ws300.iter_rows(values_only=True)
    if r[0] and str(r[0]) != "公司"
}
print(f"300 大名單 ∩ xlsx 企業名單: {len(xlsx_names & lst300)}")
print(
    f"xlsx 快照缺今日新增: {len(set(gt_by_name) - xlsx_names)}；xlsx 有但今日消失: {sorted(xlsx_names - set(gt_by_name))}"
)

json.dump(
    {
        "simulated": sim,
        "site_diffs": diffs,
        "added": added,
        "removed": removed,
        "status_changed": changed,
    },
    open(f"{SP}/comparison_result.json", "w", encoding="utf-8"),
    ensure_ascii=False,
    indent=1,
)
print("\nsaved -> comparison_result.json")
