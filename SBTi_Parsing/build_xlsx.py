# 合併重建 臺灣國際減碳倡議企業名單.xlsx（Drive id 1AyCyX5zS-oBaAe102LSAchVmPPAvHN7b）
#
# 用法：python3 build_xlsx.py <base_xlsx> <output_xlsx>
#   base_xlsx  = Drive 現版下載檔（2026-07-29 起含 qyke 的「存檔企業名單」分頁）
#   output_xlsx = 合併輸出，之後以 gws drive files update --upload 同 id 上傳
#
# 合併規則（2026-07-29 Joseph 定案，見 decisions.md）：
# - 保留 qyke 的「存檔企業名單」原樣；移除其「最新20260731企業名單」
#   （211 家內容已逐格含於新建的企業名單，零損失）
# - 新建「企業名單」於第一位：211 家取自 data/greentrade_companies.json，
#   樣式（表頭、色碼、欄寬）沿用存檔分頁（＝原版）；公司網址由存檔分頁名稱精確比對帶回
# - 統計摘要改 COUNTIF/COUNTIFS 公式；說明頁更新；2025年永續獎 B:E 改指企業名單
#   （列 5 日月光為 qyke 手填靜態值，不動）；所有排碳大戶_2024_300 不動
import json
import sys
from copy import copy

import openpyxl
from openpyxl.styles import PatternFill

SBTI_FILL = {
    "Aligned to 1.5°C": "FFC8E6C9",
    "Well-below 2°C": "FFFFF9C4",
    "Committed to set SBT target": "FFFFE0B2",
    "X": "FFFFEBEE",
}
FLAG_FILL = {"V": "FFC8E6C9", "X": "FFFFEBEE"}
AWARD_COL_MAP = {2: "D", 3: "E", 4: "F", 5: "G"}  # 永續獎 B:E → 企業名單 D:G


def fill(argb):
    return PatternFill(fill_type="solid", fgColor=argb)


def main(base_path, out_path):
    with open("data/greentrade_companies.json") as f:
        recs = json.load(f)["records"]
    assert len(recs) == 211, len(recs)

    wb = openpyxl.load_workbook(base_path)
    assert "存檔企業名單" in wb.sheetnames, wb.sheetnames
    arc = wb["存檔企業名單"]

    if "最新20260731企業名單" in wb.sheetnames:
        wb.remove(wb["最新20260731企業名單"])

    # 樣式模板與網址對照取自存檔分頁（＝原版企業名單）
    tpl = {
        c: (
            copy(arc.cell(2, c).font),
            copy(arc.cell(2, c).alignment),
            copy(arc.cell(2, c).border),
        )
        for c in range(1, 11)
    }
    url_map = {}
    for r in range(2, arc.max_row + 1):
        name, cell_i = arc.cell(r, 2).value, arc.cell(r, 9)
        if name and cell_i.value:
            url_map[str(name).strip()] = (
                cell_i.value,
                cell_i.hyperlink.target if cell_i.hyperlink else None,
            )

    ws = wb.create_sheet("企業名單", 0)
    for c in range(1, 11):
        src, dst = arc.cell(1, c), ws.cell(1, c)
        dst.value = src.value
        dst.font, dst.fill, dst.alignment, dst.border = (
            copy(src.font),
            copy(src.fill),
            copy(src.alignment),
            copy(src.border),
        )
    ws.row_dimensions[1].height = arc.row_dimensions[1].height
    for key, dim in arc.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[key].width = dim.width

    for i, rec in enumerate(recs):
        row = i + 2
        name = rec["公司名稱"].replace("臺", "台").strip()
        values = [
            i + 1,
            name,
            rec["產業別"],
            rec["SBTi"],
            rec["RE100"],
            rec["EV100"],
            rec["SEC"],
            f"=XLOOKUP(B{row},'所有排碳大戶_2024_300'!A:A,'所有排碳大戶_2024_300'!A:A)",
            url_map.get(name, ("", None))[0],
            rec["公司簡介"],
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row, c, v)
            cell.font, cell.alignment, cell.border = tpl[c]
        ws.cell(row, 4).fill = fill(SBTI_FILL[rec["SBTi"]])
        for c, key in ((5, "RE100"), (6, "EV100"), (7, "SEC")):
            ws.cell(row, c).fill = fill(FLAG_FILL[rec[key]])
        link = url_map.get(name, ("", None))[1]
        if link:
            ws.cell(row, 9).hyperlink = link
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = "A1:J212"
    carried = sum(
        1 for rec in recs if rec["公司名稱"].replace("臺", "台").strip() in url_map
    )

    st = wb["統計摘要"]
    for r in range(4, 9):
        st.cell(r, 2).value = f"=COUNTIF('企業名單'!$C:$C,$A{r})"
        st.cell(
            r, 3
        ).value = f"=COUNTIFS('企業名單'!$C:$C,$A{r},'企業名單'!$D:$D,\"<>X\")"
        st.cell(
            r, 4
        ).value = f"=COUNTIFS('企業名單'!$C:$C,$A{r},'企業名單'!$E:$E,\"V\")"
        st.cell(
            r, 5
        ).value = f"=COUNTIFS('企業名單'!$C:$C,$A{r},'企業名單'!$F:$F,\"V\")"
        st.cell(
            r, 6
        ).value = f"=COUNTIFS('企業名單'!$C:$C,$A{r},'企業名單'!$G:$G,\"V\")"
    for r in range(15, 18):
        st.cell(r, 2).value = f"=COUNTIF('企業名單'!$D:$D,A{r})"
    st.cell(18, 2).value = "=COUNTIF('企業名單'!$D:$D,\"X\")"

    sm = wb["說明"]
    sm["A6"] = "資料最後更新時間：2026/07/28（依網站標示）"
    sm["A7"] = "本檔抓取範圍：第1頁至第11頁，共計 211 筆企業"
    sm["A14"] = (
        "EP100：能源效率提升 100% —— V 表示已加入承諾（網站欄位名稱現已改為 SEC，本檔沿用原 EP100 欄名）"
    )
    sm["A21"] = (
        "1. 企業名單：完整 211 筆企業資料，含公司簡介，官網超連結沿用前版留存（本次新增企業未附），已套用篩選與凍結窗格"
    )
    sm["A23"] = "3. 存檔企業名單：前版名單（網站 2026/04/25 版，199 筆）留存"

    aw = wb["2025年永續獎"]
    for r in range(2, 40):
        if r == 5:  # 日月光：qyke 手填靜態值，greentrade 無此公司，保留
            continue
        for c, col in AWARD_COL_MAP.items():
            aw.cell(
                r, c
            ).value = f"=XLOOKUP($A{r},'企業名單'!$B:$B,'企業名單'!{col}:{col})"

    wb.save(out_path)
    print("saved:", out_path)
    print("sheets:", wb.sheetnames)
    print("URL carried over:", carried, "/ 211")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
